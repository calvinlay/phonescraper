import openpyxl
import requests
from bs4 import BeautifulSoup


def main():
    wb = openpyxl.load_workbook(r'c:testsheet.xlsx') #Add name of column
    sheet = wb['Sheet1'] #Add sheet to read/write

    for row in range(1, sheet.max_row+1): 
        if (sheet.cell(row, column=1).value is not None): #read each row from column 1 until empty cell
            cellValue = sheet.cell(row, column=1).value #read value
            number = webscrap(cellValue)
            sheet.cell(row, column=2).value = number #write value
            
    wb.save('testsheet.xlsx')

def webscrap(schoolName):
    url = "https://www.google.com/search?q=" + schoolName #website URL
    s =  requests.get(url).text #send request for url
    soup = BeautifulSoup(s, "html.parser") #scrap page
    if(soup.find("div", {"class": "AVsepf u2x1Od"}) is not None): #check if exist
        if(soup.find("span", {"class": "BNeawe tAd8D AP7Wnd"}) is not None):
            phoneNum = soup.find("div", {"class": "AVsepf u2x1Od"}).find("span", {"class": "BNeawe tAd8D AP7Wnd"}) #tag for phone number
            print(phoneNum.text)
            return str(phoneNum.text) 
    print("No Number Found")
    return "No Number Found"
    

if __name__ == "__main__":
    main()